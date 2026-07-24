import csv
import json
import openpyxl
from pathlib import Path

BASE = Path(__file__).resolve().parent / "data"
SAMPLE_ANNOT_MODEL = "comment_annotations"


def _safe_json(val):
    if not val:
        return []
    try:
        return json.loads(val)
    except (json.JSONDecodeError, TypeError):
        return []


def load_posts():
    """Return dict keyed by (post_id, model_name) with all LLM output fields."""
    wb = openpyxl.load_workbook(BASE / "PostLevel_Outputs.xlsx", read_only=True)
    ws = wb[wb.sheetnames[0]]
    rows = list(ws.iter_rows(values_only=True))
    headers = rows[0]
    posts = {}
    for row in rows[1:]:
        d = dict(zip(headers, row))
        key = (d["post_id"], d["model_name"])
        posts[key] = {
            "class_label": d["class_label"],
            "post_id": d["post_id"],
            "title": d["title"],
            "link": d["link"],
            "model_family": d["model_family"],
            "model_name": d["model_name"],
            "summary": d["summary"] or "",
            "advice": _safe_json(d["unique_advice_json"]),
            "divergences": _safe_json(d["divergences_json"]),
            "clinical_notes": _safe_json(d["clinically_relevant_notes_json"]),
            "data_quality": d["data_quality"] or "",
        }
    wb.close()
    return posts


def load_comments():
    """Return dict keyed by post_id with title, body, and list of comments."""
    wb = openpyxl.load_workbook(BASE / "6K_data_with_comments.xlsx", read_only=True)
    ws = wb[wb.sheetnames[0]]
    rows = list(ws.iter_rows(values_only=True))
    headers = rows[0]
    comment_cols = [h for h in headers if h and h.startswith("Comment")]
    data = {}
    for row in rows[1:]:
        d = dict(zip(headers, row))
        pid = d["id"]
        comments = []
        for col in comment_cols:
            val = d.get(col)
            if val and str(val).strip():
                comments.append(str(val).strip())
        data[pid] = {
            "post_id": pid,
            "title": d.get("title", ""),
            "body": d.get("body", ""),
            "label1": d.get("Label1", ""),
            "label2": d.get("Label2", ""),
            "label3": d.get("Label3", ""),
            "comments": comments,
        }
    wb.close()
    return data


def load_sample_annotations():
    """Load comment-level annotations from sample_annotations.csv.

    Returns posts dict entries keyed by (post_id, SAMPLE_ANNOT_MODEL) and
    a dict of comment data keyed by post_id (used as fallback when post
    is not in the 6K comments file).
    """
    path = BASE / "sample_annotations.csv"
    if not path.exists():
        return {}, {}

    # Group rows by post_id
    from collections import OrderedDict
    grouped = OrderedDict()
    with open(path, newline="", encoding="utf-8") as f:
        reader = csv.DictReader(f)
        for row in reader:
            pid = row["post_id"]
            if pid not in grouped:
                grouped[pid] = {
                    "title": row.get("post_title") or "",
                    "body": row.get("post_body") or "",
                    "themes": row.get("themes") or "",
                    "comments": [],
                }
            grouped[pid]["comments"].append({
                "comment_id": row.get("comment_id") or "",
                "comment_body": row.get("comment_body") or "",
                "l1_coding": row.get("l1_coding") or "",
                "span_if_claim": row.get("span_if_claim") or "",
                "nikil_notes": row.get("nikil_notes") or "",
            })

    posts = {}
    comment_fallbacks = {}
    for pid, info in grouped.items():
        # Build advice-style items from each comment annotation
        advice_items = []
        for c in info["comments"]:
            # Compute GT span offsets within the comment body
            gt_spans = []
            if c["span_if_claim"]:
                import re
                body = c["comment_body"]
                # Strip markdown formatting for matching (e.g. **NOT** → NOT)
                body_clean = re.sub(r'\*{1,2}(.+?)\*{1,2}', r'\1', body)
                # Spans are separated by commas and pipes
                raw_spans = c["span_if_claim"].replace("|", ",")
                for span_text in raw_spans.split(","):
                    span_text = span_text.strip()
                    if not span_text:
                        continue
                    # Try exact match first
                    idx = body.find(span_text)
                    if idx == -1:
                        # Try case-insensitive
                        idx = body.lower().find(span_text.lower())
                    if idx == -1:
                        # Try whitespace-normalized match (span may have single space where body has newlines)
                        span_ws = re.sub(r'\s+', ' ', span_text).strip()
                        for m in re.finditer(re.escape(span_ws[:30]), re.sub(r'\s+', ' ', body)):
                            # Found start in normalized; map back to original
                            norm_start = m.start()
                            # Walk original body to find corresponding position
                            orig_i = 0
                            norm_i = 0
                            while norm_i < norm_start and orig_i < len(body):
                                if body[orig_i].isspace():
                                    # consume all whitespace in original
                                    while orig_i < len(body) and body[orig_i].isspace():
                                        orig_i += 1
                                    norm_i += 1  # one space in normalized
                                else:
                                    orig_i += 1
                                    norm_i += 1
                            # Now verify the full span matches from orig_i
                            orig_chunk = re.sub(r'\s+', ' ', body[orig_i:orig_i + len(span_text) + 50])
                            if orig_chunk.startswith(span_ws):
                                # Find the end in original body
                                span_end = orig_i
                                ws_matched = 0
                                while span_end < len(body) and ws_matched < len(span_ws):
                                    if body[span_end].isspace():
                                        while span_end < len(body) and body[span_end].isspace():
                                            span_end += 1
                                        ws_matched += 1  # counts as one space
                                    else:
                                        span_end += 1
                                        ws_matched += 1
                                gt_spans.append({
                                    "text": body[orig_i:span_end],
                                    "start": orig_i,
                                    "end": span_end,
                                })
                                idx = orig_i  # mark as found
                                break
                    if idx == -1:
                        # Try matching against markdown-stripped version
                        idx_clean = body_clean.find(span_text)
                        if idx_clean == -1:
                            idx_clean = body_clean.lower().find(span_text.lower())
                        if idx_clean >= 0:
                            # Map offset back: find the corresponding position in original body
                            # by matching surrounding context
                            prefix = body_clean[:idx_clean]
                            # Count how many extra markdown chars precede this position
                            orig_pos = 0
                            clean_pos = 0
                            for ch in body:
                                if clean_pos >= idx_clean:
                                    break
                                if body_clean[clean_pos:clean_pos+1] == ch:
                                    clean_pos += 1
                                orig_pos += 1
                            idx = orig_pos
                            # Find the end by scanning forward
                            span_end = idx
                            clean_matched = 0
                            while span_end < len(body) and clean_matched < len(span_text):
                                if body[span_end] in '*_':
                                    span_end += 1
                                    continue
                                clean_matched += 1
                                span_end += 1
                            gt_spans.append({
                                "text": body[idx:span_end],
                                "start": idx,
                                "end": span_end,
                            })
                            continue
                    if idx >= 0:
                        gt_spans.append({
                            "text": span_text,
                            "start": idx,
                            "end": idx + len(span_text),
                        })

            advice_items.append({
                "advice": c["comment_body"],
                "comment_id": c["comment_id"],
                "agreement": c["l1_coding"],
                "support": [s.strip() for s in c["span_if_claim"].split(",") if s.strip()] if c["span_if_claim"] else [],
                "counterpoints": [c["nikil_notes"]] if c["nikil_notes"] else [],
                "gt_spans": gt_spans,
            })

        key = (pid, SAMPLE_ANNOT_MODEL)
        posts[key] = {
            "class_label": info["themes"],
            "post_id": pid,
            "title": info["title"],
            "link": f"https://www.reddit.com/r/suboxone/comments/{pid}/",
            "model_family": "sample",
            "model_name": SAMPLE_ANNOT_MODEL,
            "summary": info["body"],
            "advice": advice_items,
            "divergences": [],
            "clinical_notes": [],
            "data_quality": "",
        }

        # Fallback comment data (in case post isn't in 6K file)
        comment_fallbacks[pid] = {
            "post_id": pid,
            "title": info["title"],
            "body": info["body"],
            "label1": "",
            "label2": "",
            "label3": "",
            "comments": [c["comment_body"] for c in info["comments"]],
        }

    return posts, comment_fallbacks


def load_all():
    """Load and merge posts with their comments. Returns list of unique post_ids and full data."""
    posts = load_posts()
    comments = load_comments()

    # Load sample annotations
    sample_posts, sample_comments = load_sample_annotations()
    posts.update(sample_posts)
    # Only add comment fallbacks for posts not already in 6K data
    for pid, cdata in sample_comments.items():
        if pid not in comments:
            comments[pid] = cdata

    # Get unique post_ids preserving order
    seen = set()
    post_ids = []
    for (pid, _) in posts:
        if pid not in seen:
            seen.add(pid)
            post_ids.append(pid)

    # Get model names
    models = sorted(set(m for (_, m) in posts))

    return post_ids, posts, comments, models
