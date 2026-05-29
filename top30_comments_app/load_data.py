import json
import re
import openpyxl
import pandas as pd
from pathlib import Path

BASE = Path(__file__).resolve().parent / "data"


def _safe_json(val):
    if not val:
        return []
    try:
        return json.loads(val)
    except (json.JSONDecodeError, TypeError):
        return []


def load_posts():
    """Return dict keyed by (post_id, model_family) with all LLM output fields."""
    wb = openpyxl.load_workbook(BASE / "PostLevel_Outputs.xlsx", read_only=True)
    ws = wb[wb.sheetnames[0]]
    rows = list(ws.iter_rows(values_only=True))
    headers = rows[0]
    posts = {}
    for row in rows[1:]:
        d = dict(zip(headers, row))
        key = (str(d["post_id"]), d["model_family"])
        posts[key] = {
            "class_label": d["class_label"],
            "post_id": str(d["post_id"]),
            "title": d["title"],
            "link": d["link"],
            "model_family": d["model_family"],
            "summary": d.get("summary") or "",
            "advice": _safe_json(d.get("unique_advice_json")),
            "divergences": _safe_json(d.get("divergences_json")),
            "clinical_notes": _safe_json(d.get("clinically_relevant_notes_json")),
            "data_quality": d.get("data_quality") or "",
        }
    wb.close()
    return posts


def _parse_comments_text(text):
    """Parse 'C1: ... C2: ...' format into list of individual comments."""
    if not text or not isinstance(text, str) or not text.strip():
        return []
    # Split on C<number>: pattern
    parts = re.split(r'\n\n(?=C\d+:)', text.strip())
    comments = []
    for part in parts:
        part = part.strip()
        if not part:
            continue
        # Remove the C<n>: prefix
        cleaned = re.sub(r'^C\d+:\s*', '', part).strip()
        if cleaned:
            comments.append(cleaned)
    return comments


def load_comments():
    """Load comments from Top30 input file (has full comment text) and 6K dataset."""
    comments_data = {}

    # First load from 6K dataset (Comment1-10 columns)
    wb = openpyxl.load_workbook(BASE / "6K_data_with_comments (1).xlsx", read_only=True)
    ws = wb[wb.sheetnames[0]]
    rows = list(ws.iter_rows(values_only=True))
    headers = rows[0]
    comment_cols = [h for h in headers if h and str(h).startswith("Comment")]
    for row in rows[1:]:
        d = dict(zip(headers, row))
        pid = str(d["id"])
        body = d.get("body", "")
        if pd.isna(body):
            body = ""
        comments = []
        for col in comment_cols:
            val = d.get(col)
            if val and not pd.isna(val) and str(val).strip():
                comments.append(str(val).strip())
        comments_data[pid] = {
            "post_id": pid,
            "title": d.get("title", ""),
            "body": str(body),
            "comments": comments,
        }
    wb.close()

    # Then enrich with Top30 input file which may have post_text
    top30_path = BASE / "Top30_By_Comments_Input.xlsx"
    if top30_path.exists():
        df = pd.read_excel(top30_path, engine="openpyxl")
        for _, row in df.iterrows():
            pid = str(row.get("post_id", ""))
            if not pid:
                continue
            post_text = row.get("post_text")
            if post_text and not pd.isna(post_text) and str(post_text).strip():
                if pid in comments_data:
                    # Use post_text if body was empty
                    if not comments_data[pid]["body"].strip():
                        comments_data[pid]["body"] = str(post_text)

            # Parse comments from the combined text (may have more detail)
            comments_text = row.get("top_level_comments_text")
            if comments_text and not pd.isna(comments_text):
                parsed = _parse_comments_text(str(comments_text))
                if parsed and len(parsed) >= len(comments_data.get(pid, {}).get("comments", [])):
                    if pid in comments_data:
                        comments_data[pid]["comments"] = parsed

    return comments_data


def load_class_summaries():
    """Return dict keyed by class_label with class-level synthesis data."""
    path = BASE / "ClassLevel_Summaries.xlsx"
    if not path.exists():
        return {}
    df = pd.read_excel(path, engine="openpyxl")
    summaries = {}
    for _, row in df.iterrows():
        cl = row.get("class_label", "")
        if not cl:
            continue
        summaries[cl] = {
            "class_label": cl,
            "top_themes": _safe_json(row.get("top_themes_json")),
            "agreement_areas": _safe_json(row.get("agreement_areas_json")),
            "divergence_axes": _safe_json(row.get("divergence_axes_json")),
            "clinically_relevant_patterns": _safe_json(row.get("clinically_relevant_patterns_json")),
            "overall_takeaway": row.get("overall_takeaway") or "",
            "error": row.get("error") or "",
        }
    return summaries


def load_top30_post_ids():
    """Return set of post_ids that are in the top 30 input."""
    path = BASE / "Top30_By_Comments_Input.xlsx"
    if not path.exists():
        return set()
    df = pd.read_excel(path, engine="openpyxl")
    return set(df["post_id"].astype(str).tolist())


def load_all():
    posts = load_posts()
    comments = load_comments()
    class_summaries = load_class_summaries()
    top30_ids = load_top30_post_ids()

    # Filter to only top 30 posts
    seen = set()
    post_ids = []
    for (pid, _) in posts:
        if pid in top30_ids and pid not in seen:
            seen.add(pid)
            post_ids.append(pid)

    models = sorted(set(m for (_, m) in posts))
    return post_ids, posts, comments, models, class_summaries
