"""
Agentic MOUD peer-response analysis.

Reads Top10_High_With_Text_And_Comments.xlsx, runs your prompt on each post
(post_text + top_level_comments_text), and writes outputs to:
- Column J: OpenAI Output
- Column K: Claude Output

Uses OPENAI_API_KEY and ANTHROPIC_API_KEY from .env.

Usage:
  python run_moud_analysis.py                    # process rows missing output (both providers)
  python run_moud_analysis.py --provider claude # Claude only (fill column K)
  python run_moud_analysis.py --all              # process all 60 rows (overwrite)
  python run_moud_analysis.py --dry-run         # show what would run, no API calls
"""

from __future__ import annotations

import argparse
import os
import time
from pathlib import Path

import anthropic
import openpyxl
import openai

try:
    from dotenv import load_dotenv
    load_dotenv(Path(__file__).resolve().parent / ".env")
except ImportError:
    pass

# --- Paths ---
SCRIPT_DIR = Path(__file__).resolve().parent
EXCEL_PATH = SCRIPT_DIR / "Top10_High_With_Text_And_Comments.xlsx"

# Column indices (1-based for openpyxl)
COL_CLASS_LABEL = 1   # A
COL_POST_TEXT = 8     # H
COL_COMMENTS = 9      # I
COL_OPENAI_OUTPUT = 10  # J
COL_CLAUDE_OUTPUT = 11  # K

# Your exact prompt (unchanged)
PROMPT_TEMPLATE = """You're an expert in analysing peer response about Medications for Opioid Use Disorder (MOUD) treatment.
Below is:
The original Reddit post


All first-level comments (labeled C1, C2, etc.)


Our goal is to understand how peers give advice to manage information needs regarding MOUD treatment, and when peers agree or diverge from each other.
For this post (class_label = {class_label}):
List the UNIQUE pieces of advice given in the comments (deduplicate similar advice).


Indicate where commenters clearly agree (multiple comments reinforcing the same advice).


Indicate where commenters diverge or give conflicting advice (describe the competing views).


Note any clinically relevant issues that emerge in the advice.


Only use evidence from the comments. Do not introduce outside medical knowledge.
Return one structured output.

---
ORIGINAL POST:
{post_text}

---
FIRST-LEVEL COMMENTS:
{top_level_comments_text}

---
Provide your structured analysis:"""


def call_openai(prompt: str) -> str:
    client = openai.OpenAI(api_key=os.environ.get("OPENAI_API_KEY"))
    r = client.chat.completions.create(
        model="gpt-4o-mini",
        messages=[{"role": "user", "content": prompt}],
        temperature=0.2,
    )
    return (r.choices[0].message.content or "").strip()


def call_claude(prompt: str) -> str:
    client = anthropic.Anthropic(api_key=os.environ.get("ANTHROPIC_API_KEY"))
    r = client.messages.create(
        model="claude-sonnet-4-20250514",
        max_tokens=4096,
        messages=[{"role": "user", "content": prompt}],
        temperature=0.2,
    )
    return (r.content[0].text if r.content else "").strip()


def main():
    parser = argparse.ArgumentParser(description="Run MOUD analysis on Excel posts")
    parser.add_argument("--all", action="store_true", help="Process all rows (overwrite existing)")
    parser.add_argument("--dry-run", action="store_true", help="Show what would run, no API calls")
    parser.add_argument("--limit", type=int, default=0, help="Max rows to process (0 = no limit)")
    parser.add_argument(
        "--provider",
        choices=["openai", "claude", "both"],
        default="both",
        help="Which provider(s) to run (default: both)",
    )
    args = parser.parse_args()

    if not EXCEL_PATH.exists():
        print(f"Error: Excel not found at {EXCEL_PATH}")
        return 1

    need_openai = args.provider in ("openai", "both")
    need_claude = args.provider in ("claude", "both")
    if need_openai and not os.environ.get("OPENAI_API_KEY") and not args.dry_run:
        print("Error: OPENAI_API_KEY not set. Add it to .env or export it.")
        return 1
    if need_claude and not os.environ.get("ANTHROPIC_API_KEY") and not args.dry_run:
        print("Error: ANTHROPIC_API_KEY not set. Add it to .env or export it.")
        return 1

    wb = openpyxl.load_workbook(EXCEL_PATH, read_only=False)
    ws = wb["Sheet1"] if "Sheet1" in wb.sheetnames else wb.active

    # Ensure header for Claude column exists
    if ws.cell(row=1, column=COL_CLAUDE_OUTPUT).value is None:
        ws.cell(row=1, column=COL_CLAUDE_OUTPUT).value = "Claude Output"

    to_process = []
    for row_idx in range(2, ws.max_row + 1):
        class_label = ws.cell(row=row_idx, column=COL_CLASS_LABEL).value
        post_text = ws.cell(row=row_idx, column=COL_POST_TEXT).value
        comments = ws.cell(row=row_idx, column=COL_COMMENTS).value
        openai_exists = ws.cell(row=row_idx, column=COL_OPENAI_OUTPUT).value
        claude_exists = ws.cell(row=row_idx, column=COL_CLAUDE_OUTPUT).value

        if not post_text or not comments:
            continue
        run_openai = need_openai and (args.all or not (openai_exists and str(openai_exists).strip()))
        run_claude = need_claude and (args.all or not (claude_exists and str(claude_exists).strip()))
        if not run_openai and not run_claude:
            continue
        to_process.append((row_idx, class_label or "", post_text, comments, run_openai, run_claude))

    if args.limit:
        to_process = to_process[: args.limit]

    print(f"Rows to process: {len(to_process)} (provider: {args.provider})")
    if not to_process:
        print("Nothing to do.")
        wb.save(EXCEL_PATH)
        wb.close()
        return 0

    if args.dry_run:
        for row_idx, class_label, post_text, _, _, _ in to_process[:5]:
            print(f"  Row {row_idx}: class_label={class_label}, post_text len={len(str(post_text))}")
        if len(to_process) > 5:
            print(f"  ... and {len(to_process) - 5} more")
        wb.close()
        return 0

    for i, (row_idx, class_label, post_text, comments, run_openai, run_claude) in enumerate(to_process, 1):
        print(f"[{i}/{len(to_process)}] Row {row_idx} (class={class_label})...", end=" ", flush=True)
        prompt = PROMPT_TEMPLATE.format(
            class_label=class_label,
            post_text=post_text,
            top_level_comments_text=comments,
        )
        try:
            if run_openai:
                out = call_openai(prompt)
                ws.cell(row=row_idx, column=COL_OPENAI_OUTPUT).value = out
            if run_claude:
                out = call_claude(prompt)
                ws.cell(row=row_idx, column=COL_CLAUDE_OUTPUT).value = out
            wb.save(EXCEL_PATH)
            print("OK")
        except Exception as e:
            print(f"ERROR: {type(e).__name__}: {e}")
        time.sleep(1)  # rate limit

    wb.close()
    print("Done.")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
