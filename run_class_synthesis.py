"""
Class-level pattern synthesis for MOUD analysis.

Reads post-level LLM outputs from Top10_High_With_Text_And_Comments.xlsx, groups by class,
runs the class-level synthesis prompt on each group, and writes results to a new sheet
"ClassLevel_Synthesis" in the same Excel.

Run after run_moud_analysis.py. Processes all classes that have at least one post-level output.

Usage:
  python run_class_synthesis.py              # process all 6 classes
  python run_class_synthesis.py --dry-run    # show what would run, no API calls
  python run_class_synthesis.py --class "Access Logistics"  # single class only
"""

from __future__ import annotations

import argparse
import os
import time
from pathlib import Path

import openpyxl
import openai

try:
    from dotenv import load_dotenv
    load_dotenv(Path(__file__).resolve().parent / ".env")
except ImportError:
    pass

# --- Paths ---
SCRIPT_DIR = Path(__file__).resolve().parent
EXCEL_PATH = SCRIPT_DIR / "Top10_High_With_Text_And_Comments.xlsx"

# Column indices (1-based for openpyxl)
COL_CLASS_LABEL = 1
COL_OPENAI_OUTPUT = 10

SHEET_NAME = "ClassLevel_Synthesis"

# Your exact class-level prompt (unchanged)
CLASS_SYNTHESIS_PROMPT = """🌐 UNIVERSAL CLASS-LEVEL PATTERN SYNTHESIS PROMPT
You are analyzing multiple structured LLM outputs from posts that all belong to the same MOUD class.
Each structured output contains:
Unique advice pieces


Agreement patterns


Divergence/conflict


Clinically relevant issues


Your task is to detect patterns across the entire class.
Use only the provided structured outputs. Do not introduce outside medical knowledge.

1️⃣ Recurring Problem Pattern
Across all posts:
What type(s) of core problem recur?


Are the problems primarily:


Access-related?


Stigma-related?


Withdrawal/symptom-related?


Provider-related?


Financial?


Emotional?


Is there a dominant structural source of instability?


Summarize the main problem signature of this class.

2️⃣ Recurring Peer Strategy Pattern
Across posts:
What types of strategies are most frequently recommended?


Logistical fixes (switch provider, switch pharmacy, paperwork, etc.)


Pharmacological adjustments (dose change, formulation change, manufacturer change)


Emotional coping (ignore stigma, stay strong, validate feelings)


Confrontation/escalation (file complaint, threaten report, legal action)


Risk-taking workarounds


Which strategy type dominates?


Are strategies consistent across posts or highly variable?



3️⃣ Agreement vs Polarization Pattern
Across posts:
Is there strong consensus around certain advice?


Are there recurring ideological splits?


Do the same disagreements repeat across multiple posts?


Is advice mostly reinforcing or conflicting?


Classify the class as:
High consensus


Structured polarization


Fragmented/heterogeneous advice



4️⃣ Stability Impact Pattern
Across posts:
Do disruptions commonly threaten:


Withdrawal?


Relapse?


Treatment discontinuation?


Financial stress?


Institutional distrust?


Does peer advice appear to stabilize or destabilize treatment continuity overall?


Classify overall risk signal:
Low destabilization


Moderate destabilization


High destabilization



5️⃣ Repeated Entity / Trigger Mapping
Across posts:
Do certain named entities repeatedly appear?
 (e.g., specific pharmacies, manufacturers, insurance mechanisms, doctors, ER, etc.)


Do specific triggers repeatedly provoke conflict?


Are there repeated "villains" or repeated "trusted actors"?


Identify recurring clusters.

6️⃣ Class Signature
Provide a concise synthesis:
What defines this class?


What recurring pattern differentiates it from other MOUD classes?


What is the dominant peer narrative in this class?


---
CLASS: {class_label}

---
POST-LEVEL STRUCTURED OUTPUTS:

{post_outputs}

---
Provide your class-level synthesis:"""


def call_openai(prompt: str) -> str:
    client = openai.OpenAI(api_key=os.environ.get("OPENAI_API_KEY"))
    r = client.chat.completions.create(
        model="gpt-4o-mini",
        messages=[{"role": "user", "content": prompt}],
        temperature=0.2,
    )
    return (r.choices[0].message.content or "").strip()


def main():
    parser = argparse.ArgumentParser(description="Run class-level synthesis on MOUD post outputs")
    parser.add_argument("--dry-run", action="store_true", help="Show what would run, no API calls")
    parser.add_argument("--class", dest="single_class", type=str, help="Process only this class")
    args = parser.parse_args()

    if not EXCEL_PATH.exists():
        print(f"Error: Excel not found at {EXCEL_PATH}")
        return 1

    key = os.environ.get("OPENAI_API_KEY")
    if not key and not args.dry_run:
        print("Error: OPENAI_API_KEY not set. Add it to .env or export it.")
        return 1

    wb = openpyxl.load_workbook(EXCEL_PATH, read_only=False)
    ws = wb.active

    # Group rows by class_label, collect post-level outputs
    class_to_outputs: dict[str, list[str]] = {}
    for row_idx in range(2, ws.max_row + 1):
        class_label = ws.cell(row=row_idx, column=COL_CLASS_LABEL).value
        output = ws.cell(row=row_idx, column=COL_OPENAI_OUTPUT).value
        if not class_label:
            continue
        class_label = str(class_label).strip()
        if args.single_class and class_label != args.single_class:
            continue
        if output and str(output).strip():
            class_to_outputs.setdefault(class_label, []).append(str(output).strip())

    # Process all classes that have at least one output
    to_process = {cls: outputs for cls, outputs in class_to_outputs.items() if outputs}

    if not to_process:
        print("No classes with post-level outputs. Run run_moud_analysis.py first.")
        wb.close()
        return 1

    print(f"Classes to synthesize: {list(to_process.keys())}")

    if args.dry_run:
        for cls, outputs in to_process.items():
            total_chars = sum(len(o) for o in outputs)
            print(f"  {cls}: 10 outputs, ~{total_chars} chars total")
        wb.close()
        return 0

    # Create or get ClassLevel_Synthesis sheet
    if SHEET_NAME in wb.sheetnames:
        synth_sheet = wb[SHEET_NAME]
        synth_sheet.delete_rows(1, synth_sheet.max_row)
    else:
        synth_sheet = wb.create_sheet(SHEET_NAME)

    # Header row
    synth_sheet.cell(row=1, column=1, value="class_label")
    synth_sheet.cell(row=1, column=2, value="Class_Level_Synthesis")

    for i, (class_label, outputs) in enumerate(to_process.items(), 1):
        print(f"[{i}/{len(to_process)}] {class_label}...", end=" ", flush=True)
        post_outputs = "\n\n---\n\n".join(
            f"POST {j}:\n{out}" for j, out in enumerate(outputs, 1)
        )
        prompt = CLASS_SYNTHESIS_PROMPT.format(
            class_label=class_label,
            post_outputs=post_outputs,
        )
        try:
            synthesis = call_openai(prompt)
            row_idx = i + 1
            synth_sheet.cell(row=row_idx, column=1, value=class_label)
            synth_sheet.cell(row=row_idx, column=2, value=synthesis)
            wb.save(EXCEL_PATH)
            print("OK")
        except Exception as e:
            print(f"ERROR: {type(e).__name__}: {e}")
        time.sleep(1)

    wb.close()
    print("Done. Results in sheet 'ClassLevel_Synthesis'.")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
